#Todos los modulos que se usaran
#openpyxl es para modificar excel utilizando python
#click es para usar el comando cls desde python
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import click

#Clase producto, es el molde utilizado para cada producto, en este caso los carros
class Producto():

    #Aqui estan todos los atributos que puede tener un carro
    def __init__(self, fecha, marca, modelo, año, precio, linea, estado):
        self.fecha = fecha
        self.marca = marca
        self.modelo = modelo
        self.año = año
        self.precio = precio
        self.linea = linea
        self.estado = estado
        self.pack = [self.fecha, self.marca, self.modelo, self.año, self.precio, self.estado]
        self.data = [self.marca, self.modelo, self.año]

    def descripcion(self):
        descripcion = f"Fecha (dd/mm/aa): {self.fecha}\nMarca: {self.marca}\nModelo: {self.modelo}\nAño: {self.año}\nPrecio: {self.precio}"
        return descripcion
    
    #Esta funcion sirve para mostrar el producto en forma de tabla
    def mostrar(self):
        mostrar = f"{self.fecha}    {self.marca}    {self.modelo}     {self.año}   {self.precio}   {self.estado}"
        return mostrar

#Clase Inventario, este es como un almacen en donde encontramos todo lo que tenemos.
class Inventario():

    #Aqui se define el workbook que se va a utilizar, y las worksheets, ademas de una lista con los productos existentes.
    def __init__(self, wb):
        self.wb = wb
        self.inventario = wb['Almacenamiento']
        self.registro = wb['Vendedores']
        self.productos = []
        self.get_productos()
    
    #Esta funcion añade el producto a excel y guarda el excel para no perder el progreso
    def anadir_producto(self, producto):
        self.inventario.append(producto.pack)
        self.wb.save('Bodega.xlsx')

    #Esta funcion pasa los productos que hay en el excel a una lista en python, para poder manipularlos y analizarlos.
    def get_productos(self):
        for row in range(1, self.inventario.max_row+1):
            self.productos.append(Producto(fecha = self.inventario[get_column_letter(1) + str(row)].value,
             marca = self.inventario[get_column_letter(2) + str(row)].value,
             modelo = self.inventario[get_column_letter(3) + str(row)].value,
             año = self.inventario[get_column_letter(4) + str(row)].value,
             precio = self.inventario[get_column_letter(5) + str(row)].value,
             linea = row,
             estado = self.inventario[get_column_letter(6) + str(row)].value))

    #Muestra todos los productos que hay hasta el momento
    #Itera sobre la lista de productos y utiliza su atributo mostrar
    def mostrar_productos(self):
        for item in range(len(self.productos)):
            print(self.productos[item].mostrar())

    #Registra las ventas que se hagan
    def registrar_venta(self):

        #Se piden los datos de la venta al usuario
        print("Ingrese los siguientes datos de la venta:")
        venta = [input("Fecha: "), input("Vendedor: "), input("Marca: "), input("Modelo: "), input("Año: "), input("Precio: ")]
        datos = [venta[2],venta[3],venta[4]]
        row = 0

        #Itera por todos los productos que hay en el sistema, para analizarlos posteriormente
        for i in range(len(self.productos)):

            #Se revisa si ese carro se encuentra en el inventario, y si revisa si el carro esta a la venta
            if self.productos[i].data == datos and self.productos[i].estado == "En venta":
                row = i+1

        #Si esta a la venta se procede a registrar la venta en el excel, en caso contrario se le avisa al vendedor que el producto no esta a la venta.
        if row > 0:
            self.registro.append(venta)
            self.inventario[f'F{row}'].value = 'Vendido'
            self.productos[row-1].estado = "Vendido"
            self.wb.save("Bodega.xlsx")
            print("Venta registrada!")
        else:
            print("Este producto no esta a la venta")

#Funcion para añadir productos al excel, usando la funcion añadir_producto de la objeto inventario
def anadir_producto(inventario):
    inventario.anadir_producto(Producto(fecha = input("Fecha: "),
     marca = input("Marca: "),
     modelo = str(input("Modelo: ")),
     año = str(input("Año: ")), 
     precio = int(input("Precio: ")), 
     linea = inventario.inventario.max_row,
     estado = "En venta"))
    print("\nArticulo registrado!\n")

#Funcion para consultar los elementos en el inventario
def consultar_inventario(inventario):

    #Hay dos opciones, buscar disponibilidad de un carro, o ver todo el inventario.
    accion = input("1) Buscar disponibilidad de carro\n2) Ver inventario en general\n")
    cls()
    if accion == '1':

        #Se piden los datos del carro a buscar
        print("Ingrese los siguientes datos del carro")
        datos = [input("Marca: "), input("Modelo: "), input("Año: ")]
        count = 0

        #Se busca el carro en todos los productos, y se cuenta cuantas veces se encuentra
        for i in range(len(inventario.productos)):
            if inventario.productos[i].data == datos:
                count += 1

        #Si hay carros disponibles, se le dice cuantos hay, caso contrario se le dice que no hay.
        if count >= 1:
            print(f"Actualmente hay {count} carros en existencia")
        elif count == 0:
            print(f"Lo sentimos mucho, ese producto ya esta agotado")
    elif accion == '2':

        #Se muestran todos los productos utilizando la funcion mostrar_productos del objeto inventario
        print("  Fecha   ||  Marca  || Modelo || Año ||    Precio   ||   Estado   ||")
        inventario.mostrar_productos()
    else:
        print("Introduza 1 o 2")

#Funcion para consultar las ventas registradas en el sistema
def consultar_ventas(inventario, opcion):
    
    #Hay dos opciones, o se buscan las ventas que ha hecho un vendedor, o se busca cuantas veces se ha vendido un carro
    if opcion == "1":
        #Se pide el nombre del vendedor para buscar que carros ha vendido
        nombre = input("Ingrese el nombre del vendedor: \n")
        ventas = 0
        ganancias = 0

        #Se itera en el registro de las ventas
        for i in range(inventario.registro.max_row+1):

            #Se busca que el registro de la venta sea del vendedor que buscamos
            #Cuando se encuentre una venta del vendedor, se muestra la venta
            if inventario.registro[f"B{i+1}"].value == nombre:

                ventas += 1
                ganancias += int(inventario.registro[f"F{i+1}"].value)

                marca = inventario.registro[f"C{i+1}"].value
                modelo = inventario.registro[f"D{i+1}"].value
                año = inventario.registro[f"E{i+1}"].value
                fecha = inventario.registro[f"A{i+1}"].value
                precio = inventario.registro[f"F{i+1}"].value

                carro = f"{marca} {modelo} {año} - {fecha} {precio}"
                print(carro)

        #Se le dice al usuario cuantos carros ha vendido el vendedor y cuales han sido las ganancias
        print(f"{nombre} ha vendido {ventas} carros.")
        print(f"La ganancia total ha sido ${ganancias}")

    elif opcion == "2":

        #Se piden los datos del carro a buscar
        print("Ingrese los datos del articulo:")
        marca = input("Marca: ")
        modelo = input("Modelo: ")
        año = input("Año: ")
        datos = [marca, modelo, año]
        print("Ventas: ")
        ventas = 0
        ganancias = 0
        
        #Se itera en el registro para buscar el carro 
        for i in range(inventario.registro.max_row+1):
            datos_aux = [inventario.registro[f"C{i+1}"].value, inventario.registro[f"D{i+1}"].value, inventario.registro[f"E{i+1}"].value]
            
            #Si se encuentra el carro, se imprimen los datos de la venta
            if datos == datos_aux:
                ventas += 1
                venta = inventario.registro[f"F{i+1}"].value
                fecha = inventario.registro[f"A{i+1}"].value
                ganancias += int(venta)
                print(f"{datos_aux[0]} {datos_aux[1]} {datos_aux[2]} - {fecha} {venta}")
        
        #En caso de que haya ventas, se muestran las ventas y las ganancias, caso contrario se le dice al usuario que no hay ventas del carro
        if ventas > 0:
            print(f"Ventas: {ventas}")
            print(f"Ganancias: {ganancias}")
        else:
            print(f"No hay ventas de este articulo")
    else:
        print("ERROR")

#Funcion para registrar venta en el excel
def registrar_venta(inventario):

    #Se utiliza la funcion registrar_venta del objeto inventario
    inventario.registrar_venta()

#Funcion para limpiar la terminal
def cls():
    click.clear()

#Funcion que sirve como menu para definir que hara el usuario, y dirigirlo a la funcion necesaria
def accion(abierto, inventario):
    print("===========================================================")
    accion = input('Accion a realizar:\n1)AÑADIR PRODUCTO\n2)CONSULTAR INVENTARIO\n3)CONSULTAR VENTAS POR VENDEDOR\n4)CONSULTAR VENTAS POR ARTICULO\n5)REGISTRAR VENTA\n6)CERRAR\n')
    cls()
    if accion == '1':
        cls()
        anadir_producto(inventario)
    elif accion == '2':
        cls()
        consultar_inventario(inventario)
    elif accion == '3':
        cls()
        consultar_ventas(inventario, "1")
    elif accion == '4':
        cls()
        consultar_ventas(inventario, "2")
    elif accion == '5':
        cls()
        registrar_venta(inventario)
    elif accion == '6':
        abierto[0] = "1"
    else:
        cls()
        print('Tiene que insertar un numero\n')

#La funcion principal en donde se define el excel que se estara utilizando, y se repiten las acciones hasta que el usuario decida salir.
def main():
    abierto = ["0"]
    while abierto[0] == "0":
        wb = openpyxl.load_workbook('Bodega.xlsx')
        INVENTARIO = Inventario(wb)
        accion(abierto, INVENTARIO)
        
if __name__ == '__main__':
    main()